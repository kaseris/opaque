"""Excel report for extraction runs (spec §10).

A presentation layer over the same per-field match results the metrics aggregator produced
(§6.2) — not a second implementation of the matching logic. Three sheets: Summary,
Field-level Detail, Per-sample Rollup.
"""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..matching.results import per_sample_rollup, status_counts

if TYPE_CHECKING:
    from ..runner.runner import RunResult

_HEADER_FONT = Font(bold=True)
_SECTION_FILL = PatternFill('solid', fgColor='1F2937')
_SECTION_FONT = Font(bold=True, color='FFFFFF')
_STATUS_FILL = {
    'correct': PatternFill('solid', fgColor='C6EFCE'),
    'incorrect': PatternFill('solid', fgColor='FFC7CE'),
    'missing': PatternFill('solid', fgColor='FFEB9C'),
    'extra': PatternFill('solid', fgColor='FCE4D6'),
    'no_gold': PatternFill('solid', fgColor='D9D9D9'),
}


def build_report(result: 'RunResult', path: str | Path) -> Path:
    wb = Workbook()
    _summary_sheet(wb.active, result)
    _detail_sheet(wb.create_sheet('Field-level Detail'), result)
    _rollup_sheet(wb.create_sheet('Per-sample Rollup'), result)
    path = Path(path)
    wb.save(path)
    return path


def _summary_sheet(ws: Worksheet, result: 'RunResult') -> None:
    ws.title = 'Summary'
    counts = status_counts(result.field_results)
    m = result.metrics

    def section(title: str) -> None:
        _append_section(ws, title)

    def kv(label, value) -> None:
        ws.append([label, value])

    section('Run')
    kv('Project', result.project)
    kv('Tool', result.tool.name)
    kv('Task type', result.task_type)
    kv('Model', result.model)
    kv('Temperature', result.temperature)
    kv('Timestamp (UTC)', result.timestamp)
    kv('Git branch', result.git_branch)
    kv('Prompt bundle hash', result.prompt_bundle.bundle_hash)
    kv('Prompt dirty', result.prompt_bundle.any_dirty)

    section('Prompt versions')
    for role, fv in result.prompt_bundle.roles.items():
        kv(f'{role}', f'commit={_short(fv.git_commit)} content={fv.content_hash} dirty={fv.dirty}')

    section('Data versions')
    ev = result.eval_set_version
    kv('Eval set', f'commit={_short(ev.git_commit)} content={ev.content_hash}')
    if result.field_schema_version is not None:
        fs = result.field_schema_version
        kv('Field schema', f'commit={_short(fs.git_commit)} content={fs.content_hash}')

    section('Headline')
    kv('Overall field accuracy', m.get('field_accuracy'))
    kv('Per-sample accuracy (mean)', m.get('per_sample_accuracy_mean'))
    kv('Total samples', len(result.samples))
    kv('Labeled samples', int(m.get('labeled_sample_count', 0)))
    kv('Unlabeled samples', int(m.get('unlabeled_sample_count', 0)))
    kv('Total scored fields', int(m.get('fields_scored', 0)))

    section('Fields by status')
    for status in ('correct', 'incorrect', 'missing', 'extra', 'no_gold'):
        kv(status, counts[status])

    _autosize(ws, [34, 60])


def _detail_sheet(ws: Worksheet, result: 'RunResult') -> None:
    headers = [
        'sample_id', 'raw_file_name', 'json_path',
        'predicted_value', 'ground_truth_value', 'status', 'comparator_used',
    ]
    _write_header(ws, headers)
    status_col = headers.index('status') + 1
    for r in result.field_results:
        ws.append([
            r.sample_id,
            r.raw_file_name,
            r.json_path,
            _cell(r.predicted_value),
            _cell(r.ground_truth_value),
            r.status.value,
            r.comparator_used or '',
        ])
        fill = _STATUS_FILL.get(r.status.value)
        if fill is not None:
            ws.cell(row=ws.max_row, column=status_col).fill = fill
    ws.freeze_panes = 'A2'
    _autosize(ws, [12, 18, 34, 26, 26, 11, 15])


def _rollup_sheet(ws: Worksheet, result: 'RunResult') -> None:
    headers = ['raw_file_name', 'fields_correct', 'fields_total', 'accuracy_pct']
    _write_header(ws, headers)
    for row in per_sample_rollup(result.field_results):
        ws.append([
            row['raw_file_name'],
            row['fields_correct'],
            row['fields_total'],
            row['accuracy_pct'],
        ])
    ws.freeze_panes = 'A2'
    _autosize(ws, [22, 15, 14, 14])


# --- helpers -----------------------------------------------------------------------------

def _append_section(ws: Worksheet, title: str) -> None:
    ws.append([title])
    cell = ws.cell(row=ws.max_row, column=1)
    cell.fill = _SECTION_FILL
    cell.font = _SECTION_FONT


def _write_header(ws: Worksheet, headers: list[str]) -> None:
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal='left')


def _cell(value):
    """Render a leaf value for a cell; lists/dicts (unexpected at a leaf) become text."""
    if isinstance(value, (dict, list)):
        return str(value)
    return value


def _short(commit: str | None) -> str:
    return commit[:8] if commit else 'none'


def _autosize(ws: Worksheet, widths: list[int]) -> None:
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
