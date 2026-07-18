"""Excel report (spec §10)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from opaque.report import build_report
from opaque.runner import run

DEMO = Path(__file__).resolve().parents[1] / 'examples' / 'demo_project'


def test_report_has_three_spec_sheets(tmp_path):
    result = run(DEMO, 'invoices')
    path = build_report(result, tmp_path / 'report.xlsx')
    wb = load_workbook(path)
    assert wb.sheetnames == ['Summary', 'Field-level Detail', 'Per-sample Rollup']


def test_detail_sheet_row_per_field_result(tmp_path):
    result = run(DEMO, 'invoices')
    wb = load_workbook(build_report(result, tmp_path / 'report.xlsx'))
    detail = wb['Field-level Detail']
    headers = [c.value for c in detail[1]]
    assert headers == [
        'sample_id', 'raw_file_name', 'json_path',
        'predicted_value', 'ground_truth_value', 'status', 'comparator_used',
    ]
    # One data row per field result — including the no-gold sample's predicted fields.
    assert detail.max_row - 1 == len(result.field_results)
    statuses = {row[5].value for row in detail.iter_rows(min_row=2)}
    assert {'correct', 'incorrect', 'missing', 'extra', 'no_gold'} <= statuses


def test_rollup_sheet_row_per_sample(tmp_path):
    result = run(DEMO, 'invoices')
    wb = load_workbook(build_report(result, tmp_path / 'report.xlsx'))
    rollup = wb['Per-sample Rollup']
    assert [c.value for c in rollup[1]] == [
        'raw_file_name', 'fields_correct', 'fields_total', 'accuracy_pct'
    ]
    assert rollup.max_row - 1 == len(result.samples)
